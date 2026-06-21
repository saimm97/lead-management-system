from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.enums import ManagerType, UserRole
from app.core.security import hash_password
from app.models.lead import Lead
from app.models.profile import Profile
from app.models.user import User
from app.services.tenant import get_default_cluster_head_id


LEAD_COLUMNS = [
    "company",
    "job_title",
    "job_source",
    "primary_tech",
    "technologies",
    "phase",
    "type",
    "status",
    "notes",
    "engineer_devsinc_id",
]

PROFILE_COLUMNS = [
    "full_name",
    "linkedin_url",
    "github_url",
    "primary_tech_stack",
    "engineer_devsinc_id",
]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def _cell(row: dict[str, Any], key: str) -> str | None:
    val = row.get(key)
    if val is None:
        return None
    text = str(val).strip()
    return text or None


def _parse_rows(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_header(h) for h in rows[0]]
    parsed: list[dict[str, Any]] = []
    for raw in rows[1:]:
        if not any(raw):
            continue
        row = {headers[i]: raw[i] for i in range(min(len(headers), len(raw))) if headers[i]}
        parsed.append(row)
    return parsed


async def _resolve_engineer(db: AsyncSession, devsinc_id: str | None) -> int | None:
    if not devsinc_id:
        return None
    result = await db.execute(
        select(User).where(User.devsinc_id == devsinc_id, User.role == UserRole.ENGINEER)
    )
    engineer = result.scalar_one_or_none()
    return engineer.id if engineer else None


async def import_leads_from_excel(content: bytes, db: AsyncSession, created_by_id: int, tenant_id: int | None = None) -> dict:
    rows = _parse_rows(content)
    created = 0
    skipped = 0
    errors: list[str] = []
    default_cluster_head = await get_default_cluster_head_id(db, tenant_id)

    for idx, row in enumerate(rows, start=2):
        company = _cell(row, "company")
        job_title = _cell(row, "job_title")
        if not company or not job_title:
            skipped += 1
            errors.append(f"Row {idx}: missing company or job_title")
            continue

        technologies_raw = _cell(row, "technologies") or ""
        technologies = [t.strip() for t in technologies_raw.split(",") if t.strip()]
        engineer_id = await _resolve_engineer(db, _cell(row, "engineer_devsinc_id"))

        lead = Lead(
            tenant_id=tenant_id,
            company=company,
            job_title=job_title,
            job_source=_cell(row, "job_source") or "Import",
            primary_tech=_cell(row, "primary_tech"),
            technologies=technologies,
            phase=_cell(row, "phase") or "Applied",
            type=_cell(row, "type") or "JD Sent",
            status=_cell(row, "status") or "JD Invite Pending",
            notes=_cell(row, "notes"),
            assigned_engineer_id=engineer_id,
            cluster_head_id=default_cluster_head,
            bd_id=created_by_id,
        )
        db.add(lead)
        created += 1

    await db.flush()
    return {"created": created, "skipped": skipped, "errors": errors}


async def import_profiles_from_excel(content: bytes, db: AsyncSession, tenant_id: int | None = None) -> dict:
    rows = _parse_rows(content)
    created = 0
    skipped = 0
    errors: list[str] = []

    for idx, row in enumerate(rows, start=2):
        full_name = _cell(row, "full_name")
        if not full_name:
            skipped += 1
            errors.append(f"Row {idx}: missing full_name")
            continue

        github_url = _cell(row, "github_url")
        engineer_id = await _resolve_engineer(db, _cell(row, "engineer_devsinc_id"))

        profile = Profile(
            tenant_id=tenant_id,
            full_name=full_name,
            linkedin_url=_cell(row, "linkedin_url"),
            github_url=github_url,
            github_present=bool(github_url),
            primary_tech_stack=_cell(row, "primary_tech_stack"),
            assigned_engineer_id=engineer_id,
        )
        db.add(profile)
        created += 1

    await db.flush()
    return {"created": created, "skipped": skipped, "errors": errors}


async def import_users_from_excel(content: bytes, db: AsyncSession, tenant_id: int | None = None, manager: User | None = None) -> dict:
    rows = _parse_rows(content)
    created = 0
    skipped = 0
    errors: list[str] = []

    for idx, row in enumerate(rows, start=2):
        email = _cell(row, "email")
        full_name = _cell(row, "full_name")
        employee_id = _cell(row, "employee_id")
        role_raw = (_cell(row, "role") or "engineer").lower()
        password = _cell(row, "password") or "ChangeMe123!"

        if not email or not full_name or not employee_id:
            skipped += 1
            errors.append(f"Row {idx}: missing email, full_name, or employee_id")
            continue

        try:
            role = UserRole(role_raw)
        except ValueError:
            skipped += 1
            errors.append(f"Row {idx}: invalid role '{role_raw}'")
            continue

        if manager and manager.role == UserRole.MANAGER:
            if manager.manager_type and manager.manager_type.value == "bd_manager" and role != UserRole.BD:
                skipped += 1
                errors.append(f"Row {idx}: BD managers can only import BD team members")
                continue
            if manager.manager_type and manager.manager_type.value == "engineering_manager" and role != UserRole.ENGINEER:
                skipped += 1
                errors.append(f"Row {idx}: Engineering managers can only import engineers")
                continue

        existing = await db.execute(select(User).where((User.email == email) | (User.employee_id == employee_id)))
        if existing.scalar_one_or_none():
            skipped += 1
            errors.append(f"Row {idx}: email or employee_id already exists")
            continue

        user = User(
            tenant_id=tenant_id,
            email=email,
            password_hash=hash_password(password),
            full_name=full_name,
            employee_id=employee_id,
            devsinc_id=_cell(row, "devsinc_id") if role == UserRole.ENGINEER else None,
            role=role,
            manager_id=manager.id if manager and manager.role == UserRole.MANAGER else None,
            must_reset_password=True,
        )
        db.add(user)
        created += 1

    await db.flush()
    return {"created": created, "skipped": skipped, "errors": errors}


def build_template(entity: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    if entity == "leads":
        ws.title = "Leads"
        ws.append(LEAD_COLUMNS)
        ws.append(["Acme Corp", "Senior React Developer", "LinkedIn", "MERN", "React,Node,PostgreSQL", "Applied", "JD Sent", "JD Invite Pending", "Imported lead", "411"])
    elif entity == "profiles":
        ws.title = "Profiles"
        ws.append(PROFILE_COLUMNS)
        ws.append(["Jane Doe", "https://linkedin.com/in/janedoe", "https://github.com/janedoe", "MERN", "411"])
    elif entity == "users":
        ws.title = "Users"
        cols = ["email", "full_name", "employee_id", "role", "devsinc_id", "password"]
        ws.append(cols)
        ws.append(["new.engineer@leadpro.com", "New Engineer", "EMP999", "engineer", "412", "ChangeMe123!"])
    else:
        raise ValueError(f"Unknown entity: {entity}")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
